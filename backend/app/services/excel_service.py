"""Excel Import/Export service for Trading Log transactions."""

import io
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.transaction import Transaction

# Expected Excel columns
EXCEL_COLUMNS = ["Date", "Symbol", "Action", "Qty", "Price per Share", "Fee", "VAT", "Broker", "Note"]
VALID_ACTIONS = {"Buy", "Sell", "Snapshot"}


class ExcelExportService:
    """Service for exporting transactions to Excel."""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def export_transactions(
        self,
        user_id: uuid.UUID,
        transactions: list,
    ) -> io.BytesIO:
        """Generate an Excel file from a list of transactions.

        Returns a BytesIO buffer containing the .xlsx file.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Trading Log"

        # Header styling
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_text_font = Font(bold=True, size=11, color="FFFFFF")

        # Write headers
        for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_text_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_idx, tx in enumerate(transactions, 2):
            ws.cell(row=row_idx, column=1, value=tx.date.isoformat() if tx.date else "")
            ws.cell(row=row_idx, column=2, value=tx.stock_symbol)
            ws.cell(row=row_idx, column=3, value=tx.action)
            ws.cell(row=row_idx, column=4, value=tx.quantity)
            ws.cell(row=row_idx, column=5, value=float(tx.price_per_share))
            ws.cell(row=row_idx, column=6, value=float(tx.brokerage_fee))
            ws.cell(row=row_idx, column=7, value=float(tx.vat))
            ws.cell(row=row_idx, column=8, value=tx.broker)
            # Note from relationship
            note_text = ""
            if hasattr(tx, "note") and tx.note:
                note_text = tx.note.note if hasattr(tx.note, "note") else str(tx.note)
            ws.cell(row=row_idx, column=9, value=note_text)

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 3, 30)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


class ExcelImportService:
    """Service for importing transactions from Excel."""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def preview_import(
        self, user_id: uuid.UUID, file_bytes: bytes
    ) -> dict:
        """Validate and preview an Excel import without committing.

        Returns:
            {
                "valid_rows": int,
                "duplicate_rows": [{"row_number": int, "reason": str}],
                "error_rows": [{"row_number": int, "field": str, "error": str}],
                "preview_data": [...first 10 valid rows as dicts...],
                "total_rows": int,
            }
        """
        rows = self._parse_excel(file_bytes)
        if not rows:
            return {
                "valid_rows": 0,
                "duplicate_rows": [],
                "error_rows": [{"row_number": 1, "field": "file", "error": "No data rows found in Excel file"}],
                "preview_data": [],
                "total_rows": 0,
            }

        error_rows = []
        valid_parsed = []

        for row in rows:
            row_errors = self._validate_row(row)
            if row_errors:
                error_rows.extend(row_errors)
            else:
                valid_parsed.append(row)

        # Check duplicates for valid rows
        duplicate_rows = await self._check_duplicates(user_id, valid_parsed)
        duplicate_row_numbers = {d["row_number"] for d in duplicate_rows}

        # Rows that are valid and not duplicate
        importable = [r for r in valid_parsed if r["row_number"] not in duplicate_row_numbers]

        # Preview data (first 10 importable rows)
        preview_data = []
        for row in importable[:10]:
            preview_data.append({
                "date": row["date"].isoformat(),
                "stock_symbol": row["stock_symbol"],
                "action": row["action"],
                "quantity": row["quantity"],
                "price_per_share": float(row["price_per_share"]),
                "fee": float(row["fee"]),
                "vat": float(row["vat"]),
                "broker": row["broker"],
                "note": row.get("note", ""),
            })

        return {
            "valid_rows": len(importable),
            "duplicate_rows": duplicate_rows,
            "error_rows": error_rows,
            "preview_data": preview_data,
            "total_rows": len(rows),
        }

    async def import_transactions(
        self, user_id: uuid.UUID, file_bytes: bytes
    ) -> dict:
        """Import transactions from Excel, with full validation and duplicate check.

        Atomic: if any row fails, no rows are committed.

        Returns:
            {"imported_count": int, "message": str}
        """
        rows = self._parse_excel(file_bytes)
        if not rows:
            return {"imported_count": 0, "message": "No data rows found in Excel file"}

        error_rows = []
        valid_parsed = []

        for row in rows:
            row_errors = self._validate_row(row)
            if row_errors:
                error_rows.extend(row_errors)
            else:
                valid_parsed.append(row)

        if error_rows:
            error_summary = "; ".join(
                f"Row {e['row_number']}: {e['field']} - {e['error']}" for e in error_rows[:5]
            )
            raise ValueError(f"Validation failed: {error_summary}")

        # Check duplicates
        duplicate_rows = await self._check_duplicates(user_id, valid_parsed)
        if duplicate_rows:
            dup_summary = "; ".join(
                f"Row {d['row_number']}: {d['reason']}" for d in duplicate_rows[:5]
            )
            raise ValueError(f"Duplicate entries found: {dup_summary}")

        # All rows valid and no duplicates — create transactions
        created_count = 0
        for row in valid_parsed:
            gross_value = Decimal(str(row["quantity"])) * row["price_per_share"]
            if row["action"] == "Buy":
                net_capital_flow = gross_value + row["fee"] + row["vat"]
            else:
                net_capital_flow = gross_value - row["fee"] - row["vat"]

            transaction = Transaction(
                id=uuid.uuid4(),
                user_id=user_id,
                date=row["date"],
                stock_symbol=row["stock_symbol"],
                action=row["action"],
                quantity=row["quantity"],
                price_per_share=row["price_per_share"],
                gross_value=gross_value,
                brokerage_fee=row["fee"],
                vat=row["vat"],
                net_capital_flow=net_capital_flow,
                broker=row["broker"],
                created_at=datetime.utcnow(),
            )
            self.db.add(transaction)
            created_count += 1

        await self.db.flush()
        return {"imported_count": created_count, "message": f"Successfully imported {created_count} transactions"}

    def _parse_excel(self, file_bytes: bytes) -> list[dict]:
        """Parse Excel file and return list of row dicts with row_number."""
        try:
            wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception:
            return []

        ws = wb.active
        if ws is None:
            return []

        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if len(rows) < 2:
            return []

        # First row is header — skip it
        data_rows = []
        for idx, row in enumerate(rows[1:], start=2):
            if all(cell is None for cell in row):
                continue  # skip empty rows
            data_rows.append({
                "row_number": idx,
                "raw": row,
            })

        # Parse each row
        parsed = []
        for item in data_rows:
            row = item["raw"]
            parsed.append({
                "row_number": item["row_number"],
                "date_raw": row[0] if len(row) > 0 else None,
                "symbol_raw": row[1] if len(row) > 1 else None,
                "action_raw": row[2] if len(row) > 2 else None,
                "qty_raw": row[3] if len(row) > 3 else None,
                "price_raw": row[4] if len(row) > 4 else None,
                "fee_raw": row[5] if len(row) > 5 else None,
                "vat_raw": row[6] if len(row) > 6 else None,
                "broker_raw": row[7] if len(row) > 7 else None,
                "note_raw": row[8] if len(row) > 8 else None,
            })

        return parsed

    def _validate_row(self, row: dict) -> list[dict]:
        """Validate a single parsed row. Returns list of errors (empty if valid)."""
        errors = []
        row_num = row["row_number"]

        # Date
        parsed_date = None
        if row["date_raw"] is None:
            errors.append({"row_number": row_num, "field": "Date", "error": "Required"})
        else:
            parsed_date = self._parse_date(row["date_raw"])
            if parsed_date is None:
                errors.append({"row_number": row_num, "field": "Date", "error": "Invalid date format"})
            elif parsed_date > date.today():
                errors.append({"row_number": row_num, "field": "Date", "error": "Date cannot be in the future"})

        # Symbol
        symbol = None
        if not row["symbol_raw"]:
            errors.append({"row_number": row_num, "field": "Symbol", "error": "Required"})
        else:
            symbol = str(row["symbol_raw"]).strip().upper()
            if not symbol:
                errors.append({"row_number": row_num, "field": "Symbol", "error": "Required"})

        # Action
        action = None
        if not row["action_raw"]:
            errors.append({"row_number": row_num, "field": "Action", "error": "Required"})
        else:
            action = str(row["action_raw"]).strip().capitalize()
            if action not in VALID_ACTIONS:
                errors.append({"row_number": row_num, "field": "Action", "error": f"Must be Buy or Sell, got '{action}'"})

        # Quantity
        qty = None
        if row["qty_raw"] is None:
            errors.append({"row_number": row_num, "field": "Qty", "error": "Required"})
        else:
            try:
                qty = int(float(str(row["qty_raw"])))
                if qty <= 0:
                    errors.append({"row_number": row_num, "field": "Qty", "error": "Must be positive"})
            except (ValueError, TypeError):
                errors.append({"row_number": row_num, "field": "Qty", "error": "Must be a number"})

        # Price
        price = None
        if row["price_raw"] is None:
            errors.append({"row_number": row_num, "field": "Price", "error": "Required"})
        else:
            try:
                price = Decimal(str(row["price_raw"]))
                if price <= 0:
                    errors.append({"row_number": row_num, "field": "Price", "error": "Must be positive"})
            except (InvalidOperation, ValueError, TypeError):
                errors.append({"row_number": row_num, "field": "Price", "error": "Must be a number"})

        # Fee (optional, default 0)
        fee = Decimal("0")
        if row["fee_raw"] is not None and str(row["fee_raw"]).strip():
            try:
                fee = Decimal(str(row["fee_raw"]))
                if fee < 0:
                    errors.append({"row_number": row_num, "field": "Fee", "error": "Cannot be negative"})
            except (InvalidOperation, ValueError, TypeError):
                errors.append({"row_number": row_num, "field": "Fee", "error": "Must be a number"})

        # VAT (optional, default 0)
        vat = Decimal("0")
        if row["vat_raw"] is not None and str(row["vat_raw"]).strip():
            try:
                vat = Decimal(str(row["vat_raw"]))
                if vat < 0:
                    errors.append({"row_number": row_num, "field": "VAT", "error": "Cannot be negative"})
            except (InvalidOperation, ValueError, TypeError):
                errors.append({"row_number": row_num, "field": "VAT", "error": "Must be a number"})

        # Broker
        broker = None
        if not row["broker_raw"]:
            errors.append({"row_number": row_num, "field": "Broker", "error": "Required"})
        else:
            broker = str(row["broker_raw"]).strip()
            if not broker:
                errors.append({"row_number": row_num, "field": "Broker", "error": "Required"})

        # Note (optional)
        note = ""
        if row["note_raw"]:
            note = str(row["note_raw"]).strip()

        # If no errors, store parsed values in the row dict
        if not errors:
            row["date"] = parsed_date
            row["stock_symbol"] = symbol
            row["action"] = action
            row["quantity"] = qty
            row["price_per_share"] = price
            row["fee"] = fee
            row["vat"] = vat
            row["broker"] = broker
            row["note"] = note

        return errors

    async def _check_duplicates(
        self, user_id: uuid.UUID, valid_rows: list[dict]
    ) -> list[dict]:
        """Check for duplicate transactions.

        Duplicate = same (date, stock_symbol, action, quantity, price_per_share)
        """
        duplicates = []

        for row in valid_rows:
            stmt = select(Transaction.id).where(
                and_(
                    Transaction.user_id == user_id,
                    Transaction.date == row["date"],
                    Transaction.stock_symbol == row["stock_symbol"],
                    Transaction.action == row["action"],
                    Transaction.quantity == row["quantity"],
                    Transaction.price_per_share == row["price_per_share"],
                )
            ).limit(1)
            result = await self.db.execute(stmt)
            existing = result.scalar_one_or_none()

            if existing:
                duplicates.append({
                    "row_number": row["row_number"],
                    "reason": f"Duplicate: {row['date']} {row['stock_symbol']} {row['action']} qty={row['quantity']} price={row['price_per_share']}",
                })

        return duplicates

    @staticmethod
    def _parse_date(value) -> Optional[date]:
        """Parse a date from various formats."""
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            # Try common formats
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue
        return None
