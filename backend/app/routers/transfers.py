"""Money transfer API routes."""

import csv
import io
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import get_current_active_user
from app.models.transfer import Transfer as TransferModel
from app.models.user import User
from app.schemas.transfers import (
    TransferCreate,
    TransferFilters,
    TransferResponse,
    TransferUpdate,
)
from app.services.transfer_service import TransferService

router = APIRouter(prefix="/api/transfers", tags=["transfers"])


def _transfer_to_response(transfer) -> TransferResponse:
    """Convert a Transfer model instance to a TransferResponse."""
    return TransferResponse(
        id=str(transfer.id),
        date=transfer.date,
        broker=transfer.broker,
        transfer_type=transfer.transfer_type,
        amount=transfer.amount,
        original_currency=transfer.original_currency,
        original_amount=transfer.original_amount,
        fx_rate=transfer.fx_rate,
        converted_usd_amount=transfer.converted_usd_amount,
        fx_fee=transfer.fx_fee,
        note=transfer.note,
        created_at=transfer.created_at,
        updated_at=transfer.updated_at,
    )


@router.post("", response_model=TransferResponse, status_code=201)
async def create_transfer(
    data: TransferCreate,
    user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Create a new money transfer record."""
    service = TransferService(db)
    transfer = await service.create_transfer(user.id, data)
    return _transfer_to_response(transfer)


@router.get("/export")
async def export_transfers(
    broker: str | None = Query(None, description="Filter by broker (case-insensitive)"),
    user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Export money transfer records as CSV with all FX columns."""
    filters = TransferFilters(broker=broker)
    service = TransferService(db)
    transfers = await service.list_transfers(user.id, filters)

    # Build CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "ID", "Date", "Broker", "Transfer Type", "Amount (USD)",
        "Original Currency", "Original Amount", "FX Rate",
        "Converted USD Amount", "FX Fee", "Note",
        "FX Provider", "FX Source Timestamp", "FX Fetch Timestamp",
        "Created At", "Updated At",
    ])
    for t in transfers:
        writer.writerow([
            str(t.id),
            t.date.isoformat(),
            t.broker,
            t.transfer_type,
            str(t.amount),
            t.original_currency or "",
            str(t.original_amount) if t.original_amount is not None else "",
            str(t.fx_rate) if t.fx_rate is not None else "",
            str(t.converted_usd_amount) if t.converted_usd_amount is not None else "",
            str(t.fx_fee) if t.fx_fee is not None else "",
            t.note or "",
            t.fx_provider or "",
            t.fx_source_timestamp.isoformat() if t.fx_source_timestamp else "",
            t.fx_fetch_timestamp.isoformat() if t.fx_fetch_timestamp else "",
            t.created_at.isoformat(),
            t.updated_at.isoformat() if t.updated_at else "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=transfers_export.csv"},
    )


@router.get("", response_model=list[TransferResponse])
async def list_transfers(
    broker: str | None = Query(None, description="Filter by broker (case-insensitive)"),
    user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """List money transfer records with optional broker filter."""
    filters = TransferFilters(broker=broker)
    service = TransferService(db)
    transfers = await service.list_transfers(user.id, filters)
    return [_transfer_to_response(t) for t in transfers]


@router.put("/{transfer_id}", response_model=TransferResponse)
async def edit_transfer(
    transfer_id: uuid.UUID,
    data: TransferUpdate,
    user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Edit an existing money transfer record."""
    service = TransferService(db)
    transfer = await service.edit_transfer(user.id, transfer_id, data)
    return _transfer_to_response(transfer)


@router.delete("/{transfer_id}", status_code=204)
async def delete_transfer(
    transfer_id: uuid.UUID,
    user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Delete a money transfer record."""
    service = TransferService(db)
    await service.delete_transfer(user.id, transfer_id)


# ===== Excel Export/Import =====

TRANSFER_EXCEL_COLUMNS = ["Date", "Broker", "Type", "Amount (USD)", "Original Currency", "Original Amount", "FX Rate", "FX Fee", "Note"]
VALID_TRANSFER_TYPES = {"In", "Out"}
VALID_CURRENCIES = {"USD", "THB"}


@router.get("/export-excel")
async def export_transfers_excel(
    broker: str | None = Query(None),
    user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Export transfers as Excel (.xlsx) file."""
    filters = TransferFilters(broker=broker)
    service = TransferService(db)
    transfers = await service.list_transfers(user.id, filters)

    wb = Workbook()
    ws = wb.active
    ws.title = "Money Transfers"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, header in enumerate(TRANSFER_EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, t in enumerate(transfers, 2):
        ws.cell(row=row_idx, column=1, value=t.date.isoformat())
        ws.cell(row=row_idx, column=2, value=t.broker)
        ws.cell(row=row_idx, column=3, value=t.transfer_type)
        ws.cell(row=row_idx, column=4, value=float(t.amount))
        ws.cell(row=row_idx, column=5, value=t.original_currency or "USD")
        ws.cell(row=row_idx, column=6, value=float(t.original_amount) if t.original_amount else "")
        ws.cell(row=row_idx, column=7, value=float(t.fx_rate) if t.fx_rate else "")
        ws.cell(row=row_idx, column=8, value=float(t.fx_fee) if t.fx_fee else "")
        ws.cell(row=row_idx, column=9, value=t.note or "")

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 3, 25)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"transfers_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/import-excel/preview")
async def preview_import_transfers_excel(
    file: UploadFile = File(...),
    user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Preview and validate Excel import for transfers."""
    if not file.filename or not file.filename.endswith(".xlsx"):
        return {"error": "File must be an .xlsx Excel file"}

    file_bytes = await file.read()
    rows = _parse_transfer_excel(file_bytes)

    if not rows:
        return {"valid_rows": 0, "duplicate_rows": [], "error_rows": [{"row_number": 1, "field": "file", "error": "No data rows found"}], "preview_data": [], "total_rows": 0}

    error_rows = []
    valid_parsed = []

    for row in rows:
        errors = _validate_transfer_row(row)
        if errors:
            error_rows.extend(errors)
        else:
            valid_parsed.append(row)

    # Check duplicates
    duplicate_rows = await _check_transfer_duplicates(db, user.id, valid_parsed)
    dup_nums = {d["row_number"] for d in duplicate_rows}
    importable = [r for r in valid_parsed if r["row_number"] not in dup_nums]

    preview_data = [{"date": r["date"].isoformat(), "broker": r["broker"], "transfer_type": r["transfer_type"], "amount": float(r["amount"]), "original_currency": r["original_currency"], "note": r.get("note", "")} for r in importable[:10]]

    return {"valid_rows": len(importable), "duplicate_rows": duplicate_rows, "error_rows": error_rows, "preview_data": preview_data, "total_rows": len(rows)}


@router.post("/import-excel")
async def import_transfers_excel(
    file: UploadFile = File(...),
    user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Import transfers from Excel. Atomic: all or nothing."""
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be .xlsx")

    file_bytes = await file.read()
    rows = _parse_transfer_excel(file_bytes)
    if not rows:
        raise HTTPException(status_code=422, detail="No data rows found")

    error_rows = []
    valid_parsed = []
    for row in rows:
        errors = _validate_transfer_row(row)
        if errors:
            error_rows.extend(errors)
        else:
            valid_parsed.append(row)

    if error_rows:
        raise HTTPException(status_code=422, detail=f"Validation errors: {error_rows[:5]}")

    duplicates = await _check_transfer_duplicates(db, user.id, valid_parsed)
    if duplicates:
        raise HTTPException(status_code=422, detail=f"Duplicates found: {duplicates[:5]}")

    # Import all
    for row in valid_parsed:
        transfer = TransferModel(
            id=uuid.uuid4(),
            user_id=user.id,
            date=row["date"],
            broker=row["broker"],
            transfer_type=row["transfer_type"],
            amount=row["amount"],
            original_currency=row["original_currency"],
            original_amount=row.get("original_amount"),
            fx_rate=row.get("fx_rate"),
            converted_usd_amount=row["amount"],
            fx_fee=row.get("fx_fee"),
            note=row.get("note", ""),
            created_at=datetime.utcnow(),
        )
        db.add(transfer)

    await db.flush()
    await db.commit()
    return {"imported_count": len(valid_parsed), "message": f"Successfully imported {len(valid_parsed)} transfers"}


def _parse_transfer_excel(file_bytes: bytes) -> list[dict]:
    try:
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return []
    ws = wb.active
    if ws is None:
        return []
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(all_rows) < 2:
        return []
    parsed = []
    for idx, row in enumerate(all_rows[1:], start=2):
        if all(cell is None for cell in row):
            continue
        parsed.append({"row_number": idx, "date_raw": row[0] if len(row) > 0 else None, "broker_raw": row[1] if len(row) > 1 else None, "type_raw": row[2] if len(row) > 2 else None, "amount_raw": row[3] if len(row) > 3 else None, "currency_raw": row[4] if len(row) > 4 else None, "orig_amount_raw": row[5] if len(row) > 5 else None, "fx_rate_raw": row[6] if len(row) > 6 else None, "fx_fee_raw": row[7] if len(row) > 7 else None, "note_raw": row[8] if len(row) > 8 else None})
    return parsed


def _validate_transfer_row(row: dict) -> list[dict]:
    errors = []
    rn = row["row_number"]

    # Date
    parsed_date = _parse_date_val(row["date_raw"])
    if parsed_date is None:
        errors.append({"row_number": rn, "field": "Date", "error": "Invalid or missing date"})

    # Broker
    broker = str(row["broker_raw"]).strip() if row["broker_raw"] else ""
    if not broker:
        errors.append({"row_number": rn, "field": "Broker", "error": "Required"})

    # Type
    ttype = str(row["type_raw"]).strip().capitalize() if row["type_raw"] else ""
    if ttype not in VALID_TRANSFER_TYPES:
        errors.append({"row_number": rn, "field": "Type", "error": f"Must be In or Out, got '{ttype}'"})

    # Amount
    amount = None
    if row["amount_raw"] is not None:
        try:
            amount = Decimal(str(row["amount_raw"]))
            if amount <= 0:
                errors.append({"row_number": rn, "field": "Amount", "error": "Must be positive"})
        except (InvalidOperation, ValueError):
            errors.append({"row_number": rn, "field": "Amount", "error": "Must be a number"})
    else:
        errors.append({"row_number": rn, "field": "Amount", "error": "Required"})

    # Currency
    currency = str(row["currency_raw"]).strip().upper() if row["currency_raw"] else "USD"

    if not errors:
        row["date"] = parsed_date
        row["broker"] = broker
        row["transfer_type"] = ttype
        row["amount"] = amount
        row["original_currency"] = currency
        # Optional fields
        try:
            row["original_amount"] = Decimal(str(row["orig_amount_raw"])) if row["orig_amount_raw"] else None
        except Exception:
            row["original_amount"] = None
        try:
            row["fx_rate"] = Decimal(str(row["fx_rate_raw"])) if row["fx_rate_raw"] else None
        except Exception:
            row["fx_rate"] = None
        try:
            row["fx_fee"] = Decimal(str(row["fx_fee_raw"])) if row["fx_fee_raw"] else None
        except Exception:
            row["fx_fee"] = None
        row["note"] = str(row["note_raw"]).strip() if row["note_raw"] else ""

    return errors


async def _check_transfer_duplicates(db: AsyncSession, user_id: uuid.UUID, rows: list[dict]) -> list[dict]:
    duplicates = []
    for row in rows:
        stmt = select(TransferModel.id).where(
            and_(
                TransferModel.user_id == user_id,
                TransferModel.date == row["date"],
                TransferModel.broker == row["broker"],
                TransferModel.transfer_type == row["transfer_type"],
                TransferModel.amount == row["amount"],
            )
        ).limit(1)
        result = await db.execute(stmt)
        if result.scalar_one_or_none():
            duplicates.append({"row_number": row["row_number"], "reason": f"Duplicate: {row['date']} {row['broker']} {row['transfer_type']} ${row['amount']}"})
    return duplicates


def _parse_date_val(value) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None
